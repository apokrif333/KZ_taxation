from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from conftest_imports import SRC  # noqa: F401
from kztax270.canonical.schema import CanonicalDataset
from kztax270.canonical.workbook_schema import CANONICAL_SHEET_NAMES, required_columns
from kztax270.excel.audit_workbook import ExcelAuditWorkbookWriter, display_column_name


class WorkbookSchemaTests(unittest.TestCase):
    def test_canonical_sheet_order_is_stable(self) -> None:
        self.assertEqual(
            CANONICAL_SHEET_NAMES,
            (
                "Instruments",
                "CorporateActions",
                "Dividends",
                "Transfers",
                "Trades",
                "Fifo",
                "Positions",
                "Interest",
                "Coupons",
                "CashBalances",
                "Years_Results",
                "Unprocessed",
                "Reconciliation",
            ),
        )

    def test_reconciliation_columns_include_severity(self) -> None:
        self.assertIn("severity", required_columns("Reconciliation"))

    def test_positions_columns_do_not_include_raw_audit_noise(self) -> None:
        columns = required_columns("Positions")
        self.assertNotIn("exchange", columns)
        self.assertNotIn("calculation_price", columns)
        self.assertNotIn("source_report", columns)

    def test_dividends_columns_include_row_tax_and_exclude_old_kzt_net_tax(self) -> None:
        columns = required_columns("Dividends")
        self.assertIn("tax", columns)
        self.assertIn("tax_kzt", columns)
        self.assertNotIn("tax_kzt_usd", columns)
        self.assertNotIn("withholding_tax_kzt", columns)
        self.assertNotIn("net_amount_kzt", columns)

    def test_transfers_columns_include_fifo_price(self) -> None:
        columns = required_columns("Transfers")
        self.assertIn("quantity", columns)
        self.assertIn("price", columns)
        self.assertIn("enter_date", columns)

    def test_trades_columns_include_trade_type_marker(self) -> None:
        columns = required_columns("Trades")
        self.assertIn("trade_type", columns)

    def test_cash_balances_include_source_account(self) -> None:
        columns = required_columns("CashBalances")
        self.assertEqual(columns[:2], ("broker", "account_id"))

    def test_yearly_interest_columns_use_only_profit_tax_base(self) -> None:
        from kztax270.canonical.workbook_schema import YEARS_RESULTS_TABLE_COLUMNS

        self.assertEqual(
            YEARS_RESULTS_TABLE_COLUMNS["Yearly Interest"],
            ("year", "flag", "currency", "amount", "amount_kzt", "only_profit", "only_profit_kzt", "tax_kzt"),
        )

    def test_yearly_coupon_columns_use_only_profit_tax_base(self) -> None:
        from kztax270.canonical.workbook_schema import YEARS_RESULTS_TABLE_COLUMNS

        self.assertEqual(
            YEARS_RESULTS_TABLE_COLUMNS["Yearly Coupons"],
            (
                "year",
                "flag",
                "currency",
                "amount",
                "amount_kzt",
                "only_profit",
                "only_profit_kzt",
                "withhold_kzt",
                "tax_kzt",
                "tax_kzt_withhold",
            ),
        )

    def test_coupon_rows_preserve_explicit_revert_marker(self) -> None:
        self.assertIn("is_revert", required_columns("Coupons"))

    def test_yearly_trade_columns_use_tax_exchange_name(self) -> None:
        from kztax270.canonical.workbook_schema import YEARS_RESULTS_TABLE_COLUMNS

        columns = YEARS_RESULTS_TABLE_COLUMNS["Yearly Trades"]
        self.assertIn("tax_exchange", columns)
        self.assertNotIn("exchange", columns)
        self.assertEqual(columns[:4], ("year", "flag", "country", "tax_exchange"))
        self.assertEqual(display_column_name("tax_exchange"), "Tax_Exchange")

    def test_excel_headers_use_legacy_style_title_case(self) -> None:
        self.assertEqual(display_column_name("date_time"), "Date_Time")
        self.assertEqual(display_column_name("gross_amount_kzt"), "Gross_Amount_KZT")
        self.assertEqual(display_column_name("tax"), "Tax")
        self.assertEqual(display_column_name("only_profit_kzt"), "OnlyProfit_KZT")
        self.assertEqual(display_column_name("pnl_kzt"), "PnL_KZT")
        self.assertEqual(display_column_name("source_trade_id"), "Source_Trade_ID")

    def test_excel_writer_writes_numeric_columns_as_numbers(self) -> None:
        from openpyxl import load_workbook

        dataset = CanonicalDataset.empty("ib", "UTEST")
        dataset.tables["Trades"] = [
            {
                "date_time": "2024-01-01 10:00:00",
                "trade_id": "trade-1",
                "trade_type": "trade",
                "symbol": "AAPL",
                "isin": "US0378331005",
                "asset_type": "Stocks",
                "quantity": "10",
                "price": "100.25",
                "multiplier": "1",
                "amount": "1002.50",
                "commission": "1.25",
                "amount_with_commission": "1003.75",
                "currency": "USD",
                "country": "US",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "audit.xlsx"
            ExcelAuditWorkbookWriter().write(dataset, path)
            workbook = load_workbook(path, data_only=True)

        ws = workbook["Trades"]
        headers = [cell.value for cell in ws[1]]
        quantity_cell = ws.cell(row=2, column=headers.index("Quantity") + 1)
        amount_cell = ws.cell(row=2, column=headers.index("Amount") + 1)
        isin_cell = ws.cell(row=2, column=headers.index("ISIN") + 1)

        self.assertEqual(quantity_cell.data_type, "n")
        self.assertEqual(amount_cell.data_type, "n")
        self.assertEqual(isin_cell.data_type, "s")

    def test_years_results_blocks_have_titles_headers_and_dimension_cells_formatted(self) -> None:
        from openpyxl import load_workbook

        dataset = CanonicalDataset.empty("ib", "UTEST")
        dataset.tables["Years_Results"] = [
            {
                "table": "Yearly Interest",
                "year": "2024",
                "flag": "non-preferential",
                "currency": "USD",
                "amount": "10",
                "amount_kzt": "4690",
                "only_profit": "10",
                "only_profit_kzt": "4690",
                "tax_kzt": "469",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "audit.xlsx"
            ExcelAuditWorkbookWriter().write(dataset, path)
            workbook = load_workbook(path)

        ws = workbook["Years_Results"]
        self.assertIn("A1:H1", {str(cell_range) for cell_range in ws.merged_cells.ranges})
        self.assertEqual(ws["A1"].value, "Yearly Interest")
        self.assertTrue(ws["A1"].font.bold)
        self.assertEqual(ws["A1"].alignment.horizontal, "center")
        self.assertEqual(ws["A1"].border.bottom.style, "thin")
        self.assertTrue(all(cell.font.bold for cell in ws[2]))
        self.assertTrue(all(cell.border.bottom.style == "thin" for cell in ws[2]))
        self.assertTrue(ws["A3"].font.bold)  # Year
        self.assertTrue(ws["B3"].font.bold)  # Flag
        self.assertTrue(ws["C3"].font.bold)  # Currency
        self.assertEqual(ws["A3"].border.left.style, "thin")
        self.assertFalse(ws["D3"].font.bold)  # Amount
        self.assertEqual(ws["D3"].number_format, "0.00")
        self.assertEqual(ws["E3"].number_format, '#,##0.00 "₸"')


if __name__ == "__main__":
    unittest.main()
