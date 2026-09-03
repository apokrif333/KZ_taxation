"""Native parser for Alatay (Alatau City Invest) security and cash reports."""

from __future__ import annotations

import csv
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from kztax270.canonical.schema import AccountMetadata, CanonicalDataset
from kztax270.canonical.trade_enrichment import enrich_trades_before_calculations
from kztax270.reconciliation.models import ReconciliationMetric
from kztax270.reference.fx import AnnualFxRateProvider

from .base import BrokerReport, ParseResult
from .discovery import DiscoveryRule, discover_raw_reports
from .ib import (
    _build_broker_trade_realized_pl,
    _build_fifo_and_positions,
    _build_unprocessed_rows,
    _build_years_results,
    _canonical_trade_rows,
    _instrument_identity_key_from_values,
    _instrument_symbol_history,
    _sort_trades_by_datetime,
)

BROKER_CODE = "alatay"
RAW_FOLDER = "alatay"
BASE_CURRENCY = "KZT"

POSITIONS_SECTION = "Ценные бумаги в портфеле клиента на конец отчетного периода:"
OPENING_POSITIONS_SECTION = "Ценные бумаги в портфеле клиента на начало отчетного периода:"
TRADES_SECTION = "Движение Ценных бумаг клиента за отчетный период:"
HISTORICAL_TRADES_SECTION = (
    "Сделки по ценным бумагам, указанным на начало отчетного периода, "
    "за время до начала отчетного периода:"
)
CASH_MOVEMENTS_HEADER = "Дата проведения операции/сделки"


@dataclass(slots=True)
class ParsedAlatayReport:
    path: Path
    account_id: str | None = None
    holder_name: str | None = None
    period_start: date | None = None
    period_end: date | None = None
    positions: list[dict[str, Any]] = field(default_factory=list)
    trades: list[dict[str, Any]] = field(default_factory=list)
    security_transfers: list[dict[str, Any]] = field(default_factory=list)
    cash_movements: list[dict[str, Any]] = field(default_factory=list)
    opening_cash: Decimal | None = None
    ending_cash: Decimal | None = None
    cash_currency: str | None = None
    unprocessed: list[dict[str, Any]] = field(default_factory=list)


class AlatayParser:
    broker_code = BROKER_CODE

    def __init__(self, fx_provider: AnnualFxRateProvider | None = None) -> None:
        self.fx_provider = fx_provider or AnnualFxRateProvider({})

    def discover_reports(self, raw_root: Path, account_id: str) -> list[BrokerReport]:
        return discover_raw_reports(
            raw_root,
            DiscoveryRule(
                broker=RAW_FOLDER,
                account_id=account_id,
                extensions=frozenset({".csv", ".xlsx"}),
            ),
        )

    def parse_reports(self, reports: Sequence[BrokerReport], account_id: str) -> ParseResult:
        parsed_reports = [parse_alatay_report(report.path) for report in reports]
        dataset = build_canonical_dataset(parsed_reports, account_id, self.fx_provider)
        dataset.raw_totals.source_reports = [str(report.path) for report in reports]
        return ParseResult(
            broker=self.broker_code,
            account_id=account_id,
            reports=reports,
            dataset=dataset,
            raw_totals=dataset.raw_totals,
        )


def parse_alatay_csv(path: Path) -> ParsedAlatayReport:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return _parse_alatay_rows(path, csv.reader(handle))


def parse_alatay_xlsx(path: Path) -> ParsedAlatayReport:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        return _parse_alatay_rows(path, worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def parse_alatay_report(path: Path) -> ParsedAlatayReport:
    if path.suffix.lower() == ".csv":
        return parse_alatay_csv(path)
    if path.suffix.lower() == ".xlsx":
        return parse_alatay_xlsx(path)
    raise ValueError(f"Unsupported Alatay report format: {path}")


def _parse_alatay_rows(path: Path, rows: Sequence[Sequence[Any]] | Any) -> ParsedAlatayReport:
    parsed = ParsedAlatayReport(path=path)
    section: str | None = None

    for source_row, row in enumerate(rows, start=1):
        values = [_clean_text(value) for value in row]
        first = values[0] if values else ""
        line_text = ",".join(values).rstrip(",")

        holder_name = _metadata_value(values, "ФИО/Наименование клиента:")
        if holder_name is not None:
            parsed.holder_name = holder_name
            continue
        report_account_id = _metadata_value(values, "№ лицевого счета:")
        if report_account_id is not None:
            parsed.account_id = report_account_id
            continue
        period_text = _metadata_value(values, "Отчет составлен на:")
        if period_text is not None:
            parsed.period_start, parsed.period_end = _parse_period(period_text)
            continue
        opening_cash = _metadata_value(values, "Входящий остаток на начало периода:")
        if opening_cash is not None:
            parsed.opening_cash, parsed.cash_currency = _parse_amount_currency(opening_cash)
            continue
        ending_cash = _metadata_value(values, "Исходящий остаток на конец периода:")
        if ending_cash is not None:
            parsed.ending_cash, parsed.cash_currency = _parse_amount_currency(ending_cash)
            continue

        if first == POSITIONS_SECTION:
            section = "closing_positions"
            continue
        if first == OPENING_POSITIONS_SECTION:
            section = "opening_positions"
            continue
        if (
            first == TRADES_SECTION
            or first == "Сделки по ценным бумагам"
            or line_text == HISTORICAL_TRADES_SECTION
            or line_text.startswith("Сделки по ценным бумагам, указанным на начало отчетного периода")
        ):
            section = "trades"
            continue
        if first == CASH_MOVEMENTS_HEADER:
            section = "cash_movements"
            continue
        if not any(values):
            section = None
            continue

        if section in {"closing_positions", "opening_positions"}:
            position = _parse_position_row(
                values,
                source_report=str(path),
                source_row=source_row,
                snapshot_kind="opening" if section == "opening_positions" else "closing",
                snapshot_date=parsed.period_end,
            )
            if position is not None:
                parsed.positions.append(position)
            continue
        if section == "trades":
            trade = _parse_trade_row(values, source_report=str(path), source_row=source_row)
            if trade is None:
                continue
            if trade.get("_recognized_operation"):
                parsed.trades.append(trade)
            elif _is_internal_security_transfer(trade.get("operation")):
                parsed.security_transfers.append(trade)
            else:
                parsed.unprocessed.append(_unprocessed_trade(trade))
            continue
        if section == "cash_movements":
            movement = _parse_cash_movement_row(values, source_report=str(path), source_row=source_row)
            if movement is not None:
                parsed.cash_movements.append(movement)

    return parsed


def build_canonical_dataset(
    reports: Sequence[ParsedAlatayReport],
    account_id: str,
    fx_provider: AnnualFxRateProvider,
) -> CanonicalDataset:
    dataset = CanonicalDataset(
        metadata=AccountMetadata(broker=BROKER_CODE, account_id=account_id, base_currency=BASE_CURRENCY)
    )
    for report in reports:
        if report.account_id and not _same_account_id(report.account_id, account_id):
            dataset.warnings.append(
                f"Alatay report {report.path} belongs to account {report.account_id}, expected {account_id}."
            )

    instruments = _build_instruments(reports, account_id)
    dataset.tables["Instruments"] = instruments
    internal_trades = _sort_trades_by_datetime(_build_trades(reports, instruments))
    enrich_trades_before_calculations(dataset, internal_trades, fx_provider)
    dataset.tables["Trades"] = _canonical_trade_rows(internal_trades)
    dataset.tables["_BrokerTradeRealizedPL"] = _build_broker_trade_realized_pl(internal_trades)

    fifo_rows, positions, transfer_rows = _build_fifo_and_positions(
        internal_trades,
        transfers=[],
        initial_lots=[],
        max_year=_max_report_year(reports),
        fx_provider=fx_provider,
        warnings=dataset.warnings,
        symbol_history=_instrument_symbol_history(instruments),
    )
    dataset.tables["Fifo"] = fifo_rows
    dataset.tables["Positions"] = positions
    dataset.tables["Transfers"] = [*transfer_rows, *_build_security_transfers(reports, instruments)]
    dataset.tables["CorporateActions"] = _build_corporate_actions(reports, instruments)
    dataset.tables["Dividends"] = _build_dividends(reports, instruments, fx_provider, dataset.warnings)
    dataset.tables["Interest"] = []
    dataset.tables["Coupons"] = _build_coupons(reports, instruments, fx_provider, dataset.warnings)
    dataset.tables["CashBalances"] = _build_cash_balances(
        reports,
        account_id,
        fx_provider,
        dataset.warnings,
    )
    dataset.tables["Unprocessed"] = [
        *_build_unprocessed_rows(dataset.tables["Trades"], fifo_rows),
        *(row for report in reports for row in report.unprocessed),
    ]
    dataset.tables["Years_Results"] = _build_years_results(dataset)
    _populate_raw_totals(dataset, reports, internal_trades)
    return dataset


def _parse_position_row(
    values: Sequence[str],
    *,
    source_report: str,
    source_row: int,
    snapshot_kind: str = "closing",
    snapshot_date: date | None = None,
) -> dict[str, Any] | None:
    if len(values) < 6 or not values[0].isdigit():
        return None
    isin = values[3]
    if not isin:
        return None
    return {
        "issuer": values[1],
        "security_type": values[2],
        "isin": isin,
        "nominal": str(_decimal(values[4])),
        "quantity": str(_decimal(values[5])),
        "issuer_country": values[6] if len(values) > 6 else None,
        "source_report": source_report,
        "source_row": source_row,
        "_snapshot_kind": snapshot_kind,
        "_snapshot_date": snapshot_date.isoformat() if snapshot_date else None,
    }


def _parse_trade_row(
    values: Sequence[str],
    *,
    source_report: str,
    source_row: int,
) -> dict[str, Any] | None:
    if len(values) < 10:
        return None
    trade_date = _parse_date(values[0])
    if trade_date is None:
        return None
    operation = values[5]
    operation_key = operation.casefold()
    recognized = operation_key.startswith("покупка") or operation_key.startswith("продажа")
    quantity = _decimal(values[6])
    if operation_key.startswith("продажа"):
        quantity = -abs(quantity)
    elif operation_key.startswith("покупка"):
        quantity = abs(quantity)
    commission_index = 12 if len(values) >= 13 else 11
    issuer_country = values[11] if len(values) >= 13 else None
    return {
        "date_time": datetime.combine(trade_date, datetime.min.time()).isoformat(sep=" "),
        "issuer": values[1],
        "security_type": values[2],
        "isin": values[3],
        "nominal": str(_decimal(values[4])),
        "operation": operation,
        "quantity": str(quantity),
        "price": str(_decimal(values[7])),
        "currency": values[8] or BASE_CURRENCY,
        "amount": str(abs(_decimal(values[9]))),
        "exchange": values[10] if len(values) > 10 else None,
        "issuer_country": issuer_country,
        "commission": str(abs(_decimal(values[commission_index] if len(values) > commission_index else None))),
        "source_report": source_report,
        "source_row": source_row,
        "_recognized_operation": recognized,
    }


def _parse_cash_movement_row(
    values: Sequence[str],
    *,
    source_report: str,
    source_row: int,
) -> dict[str, Any] | None:
    if len(values) < 10:
        return None
    movement_date = _parse_date(values[0])
    if movement_date is None:
        return None
    return {
        "date": movement_date.isoformat(),
        "description": values[1],
        "operation": values[2],
        "issuer": values[3],
        "isin": values[4],
        "opening_cash": str(_decimal(values[5])),
        "credit": str(_decimal(values[6])),
        "debit": str(_decimal(values[7])),
        "ending_cash": str(_decimal(values[8])),
        "currency": values[9] or BASE_CURRENCY,
        "security_type": values[10] if len(values) > 10 else None,
        "issuer_country": values[11] if len(values) > 11 else None,
        "exchange": values[12] if len(values) > 12 else None,
        "source_report": source_report,
        "source_row": source_row,
    }


def _build_instruments(
    reports: Sequence[ParsedAlatayReport],
    account_id: str,
) -> list[dict[str, Any]]:
    sources: dict[str, dict[str, Any]] = {}
    latest_dates: dict[str, date | None] = {}
    for report in reports:
        # Trade rows carry the exchange while position rows do not, so prefer
        # them as the instrument source when both are available.
        for row in [*report.trades, *report.positions, *report.cash_movements]:
            isin = _text(row.get("isin"))
            if not isin:
                continue
            sources.setdefault(isin, row)
            current = latest_dates.get(isin)
            if report.period_end and (current is None or report.period_end > current):
                latest_dates[isin] = report.period_end

    instruments: list[dict[str, Any]] = []
    for isin in sorted(sources):
        source = sources[isin]
        country = _country_from_values(source.get("issuer_country"), isin)
        instruments.append(
            {
                "symbol": isin,
                "description": _text(source.get("issuer")) or isin,
                "conid": None,
                "security_id": isin,
                "underlying": None,
                "listing_exchange": _normalized_exchange(source.get("exchange")),
                "multiplier": "1",
                "type": _asset_type(source.get("security_type")),
                "code": None,
                "year": None,
                "expiry": None,
                "delivery_month": None,
                "strike": None,
                "issuer": _text(source.get("issuer")),
                "maturity": None,
                "cusip": None,
                "country": country,
                "isin": isin,
                "figi": None,
                "issuer_country": country,
                "offshore_flag": False if country == "KZ" else None,
                "issuer_outside_kz_flag": False if country == "KZ" else (True if country else None),
                "preferential_tax_flag": None,
                "source_broker": BROKER_CODE,
                "source_account": account_id,
                "source_report": source.get("source_report"),
                "as_of_date": latest_dates.get(isin).isoformat() if latest_dates.get(isin) else None,
            }
        )
    return instruments


def _build_trades(
    reports: Sequence[ParsedAlatayReport],
    instruments: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    instrument_lookup = {str(row.get("isin")): row for row in instruments}
    trades: list[dict[str, Any]] = []
    for report in reports:
        for row in report.trades:
            isin = _text(row.get("isin"))
            instrument = instrument_lookup.get(isin or "", {})
            quantity = _decimal(row.get("quantity"))
            price = _decimal(row.get("price"))
            amount = abs(_decimal(row.get("amount"))) or abs(quantity * price)
            commission = abs(_decimal(row.get("commission")))
            symbol = _text(instrument.get("symbol")) or isin
            country = _text(instrument.get("country")) or _country_from_values(row.get("issuer_country"), isin)
            trades.append(
                {
                    "date_time": row.get("date_time"),
                    "trade_id": f"{report.path.name}:{row.get('source_row')}",
                    "trade_type": "trade",
                    "symbol": symbol,
                    "isin": isin,
                    "asset_type": _text(instrument.get("type")) or _asset_type(row.get("security_type")),
                    "quantity": str(quantity),
                    "calculation_quantity": str(quantity),
                    "price": str(price),
                    "calculation_price": str(price),
                    "multiplier": "1",
                    "_calculation_multiplier": "1",
                    "amount": str(amount),
                    "commission": str(commission),
                    "amount_with_commission": str(amount + commission),
                    "currency": _text(row.get("currency")) or BASE_CURRENCY,
                    "exchange": _normalized_exchange(row.get("exchange")),
                    "country": country,
                    "source_report": row.get("source_report"),
                    "_instrument_identity_key": _instrument_identity_key_from_values(isin=isin, symbol=symbol),
                    "_broker_realized_pl": "0",
                }
            )
    return trades


def _build_security_transfers(
    reports: Sequence[ParsedAlatayReport],
    instruments: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    instrument_lookup = {str(row.get("isin")): row for row in instruments}
    transfers: list[dict[str, Any]] = []
    for report in reports:
        for row in report.security_transfers:
            operation = _text(row.get("operation")) or ""
            operation_key = operation.casefold()
            direction = "in" if "получатель" in operation_key else "out"
            isin = _text(row.get("isin"))
            instrument = instrument_lookup.get(isin or "", {})
            transfers.append(
                {
                    "date": _date_from_datetime(row.get("date_time")),
                    "transfer_type": "security",
                    "direction": direction,
                    "asset_type": _text(instrument.get("type")) or _asset_type(row.get("security_type")),
                    "symbol": _text(instrument.get("symbol")) or isin,
                    "isin": isin,
                    "currency": _text(row.get("currency")) or BASE_CURRENCY,
                    "quantity": str(abs(_decimal(row.get("quantity")))),
                    "price": str(_decimal(row.get("price"))),
                    "enter_date": None,
                    # This is an internal depository movement without a change
                    # of ownership, not a cash deposit/withdrawal.
                    "amount": "0",
                    "broker_comment": operation,
                    "counterparty": "internal depository transfer (no change of ownership)",
                    "source_report": row.get("source_report"),
                }
            )
    return transfers


def _build_dividends(
    reports: Sequence[ParsedAlatayReport],
    instruments: Sequence[Mapping[str, Any]],
    fx_provider: AnnualFxRateProvider,
    warnings: list[str],
) -> list[dict[str, Any]]:
    instrument_lookup = {str(row.get("isin")): row for row in instruments}
    dividends: list[dict[str, Any]] = []
    for report in reports:
        for row in report.cash_movements:
            if "дивиденд" not in str(row.get("operation") or "").casefold():
                continue
            isin = _text(row.get("isin"))
            instrument = instrument_lookup.get(isin or "", {})
            currency = _text(row.get("currency")) or BASE_CURRENCY
            gross = _decimal(row.get("credit")) - _decimal(row.get("debit"))
            year = _year_from_date(row.get("date"))
            rate = _annual_rate(fx_provider, year, currency, warnings)
            country = _text(instrument.get("country")) or _country_from_values(row.get("issuer_country"), isin)
            tax = Decimal("0") if country == "KZ" else max(gross, Decimal("0")) * Decimal("0.10")
            dividends.append(
                {
                    "date": row.get("date"),
                    "pay_date": row.get("date"),
                    "symbol": _text(instrument.get("symbol")) or isin,
                    "isin": isin,
                    "country": country,
                    "currency": currency,
                    "gross_amount": _money_text(gross),
                    "withholding_tax": "0.00",
                    "net_amount": _money_text(gross),
                    "kzt_rate": str(rate) if rate is not None else None,
                    "gross_amount_kzt": _amount_kzt(gross, rate),
                    "tax": _money_text(tax),
                    "tax_kzt": _amount_kzt(tax, rate),
                    "offshore_flag": False if country == "KZ" else None,
                    "kase_aix_preferential_flag": True if country == "KZ" else None,
                    "source_report": row.get("source_report"),
                }
            )
    return dividends


def _build_coupons(
    reports: Sequence[ParsedAlatayReport],
    instruments: Sequence[Mapping[str, Any]],
    fx_provider: AnnualFxRateProvider,
    warnings: list[str],
) -> list[dict[str, Any]]:
    instrument_lookup = {str(row.get("isin")): row for row in instruments}
    coupons: list[dict[str, Any]] = []
    for report in reports:
        for row in report.cash_movements:
            if "купон" not in str(row.get("operation") or "").casefold():
                continue
            isin = _text(row.get("isin"))
            instrument = instrument_lookup.get(isin or "", {})
            currency = _text(row.get("currency")) or BASE_CURRENCY
            gross = _decimal(row.get("credit")) - _decimal(row.get("debit"))
            year = _year_from_date(row.get("date"))
            rate = _annual_rate(fx_provider, year, currency, warnings)
            country = _text(instrument.get("country")) or _country_from_values(row.get("issuer_country"), isin)
            coupons.append(
                {
                    "date": row.get("date"),
                    "symbol": _text(instrument.get("symbol")) or isin,
                    "isin": isin,
                    "country": country,
                    "currency": currency,
                    "gross_amount": _money_text(gross),
                    "withholding_tax": "0.00",
                    "net_amount": _money_text(gross),
                    "kzt_rate": str(rate) if rate is not None else None,
                    "gross_amount_kzt": _amount_kzt(gross, rate),
                    "withholding_tax_kzt": "0.00",
                    "net_amount_kzt": _amount_kzt(gross, rate),
                    "is_revert": False,
                    "offshore_flag": False if country == "KZ" else None,
                    "source_report": row.get("source_report"),
                }
            )
    return coupons


def _build_cash_balances(
    reports: Sequence[ParsedAlatayReport],
    account_id: str,
    fx_provider: AnnualFxRateProvider,
    warnings: list[str],
) -> list[dict[str, Any]]:
    balances: list[dict[str, Any]] = []
    for report in reports:
        if report.ending_cash is None or report.period_end is None:
            continue
        currency = report.cash_currency or BASE_CURRENCY
        rate = _annual_rate(fx_provider, report.period_end.year, currency, warnings)
        balances.append(
            {
                "broker": BROKER_CODE,
                "account_id": account_id,
                "year": report.period_end.year,
                "date": report.period_end.isoformat(),
                "currency": currency,
                "ending_cash": _money_text(report.ending_cash),
                "ending_cash_kzt": _amount_kzt(report.ending_cash, rate),
                "source_report": str(report.path),
            }
        )
    return balances


def _build_corporate_actions(
    reports: Sequence[ParsedAlatayReport],
    instruments: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    instrument_lookup = {str(row.get("isin")): row for row in instruments}
    actions: list[dict[str, Any]] = []
    for report in reports:
        for row in report.cash_movements:
            operation = _text(row.get("operation")) or ""
            if "погашение цб" not in operation.casefold():
                continue
            isin = _text(row.get("isin"))
            instrument = instrument_lookup.get(isin or "", {})
            proceeds = _decimal(row.get("credit")) - _decimal(row.get("debit"))
            actions.append(
                {
                    "date": row.get("date"),
                    "symbol": _text(instrument.get("symbol")) or isin,
                    "isin": isin,
                    "action_type": "redemption",
                    "description": operation,
                    "quantity": "0",
                    "proceeds": _money_text(proceeds),
                    "value": _money_text(proceeds),
                    "currency": _text(row.get("currency")) or BASE_CURRENCY,
                    # The cash report contains principal proceeds but no lot
                    # quantity or acquisition cost from which P/L can be made.
                    "realized_pl": "0",
                    "source_report": row.get("source_report"),
                }
            )
    return actions


def _populate_raw_totals(
    dataset: CanonicalDataset,
    reports: Sequence[ParsedAlatayReport],
    trades: Sequence[Mapping[str, Any]],
) -> None:
    total_amount = Decimal("0")
    total_commission = Decimal("0")
    turnover_metric = ReconciliationMetric.TRADE_GROSS_AMOUNT_BY_INSTRUMENT.value
    for trade in trades:
        amount = abs(_decimal(trade.get("amount")))
        commission = abs(_decimal(trade.get("commission")))
        total_amount += amount
        total_commission += commission
        trade_year = _year_from_datetime(trade.get("date_time"))
        key = _dimension_key(
            metric=turnover_metric,
            year=trade_year,
            currency=_text(trade.get("currency")),
            instrument_key=_text(trade.get("isin") or trade.get("symbol")),
        )
        dataset.raw_totals.totals_by_metric_currency[key] = (
            dataset.raw_totals.totals_by_metric_currency.get(key, Decimal("0")) + amount
        )
    dataset.raw_totals.scalar_totals[ReconciliationMetric.TOTAL_TRADES_GROSS_AMOUNT.value] = total_amount
    dataset.raw_totals.scalar_totals[ReconciliationMetric.TOTAL_COMMISSIONS.value] = total_commission
    dividends_gross = sum(
        (_decimal(row.get("gross_amount")) for row in dataset.tables.get("Dividends", [])),
        Decimal("0"),
    )
    dividends_tax = sum(
        (_decimal(row.get("withholding_tax")) for row in dataset.tables.get("Dividends", [])),
        Decimal("0"),
    )
    coupon_total = sum(
        (_decimal(row.get("gross_amount")) for row in dataset.tables.get("Coupons", [])),
        Decimal("0"),
    )
    dataset.raw_totals.scalar_totals.update(
        {
            ReconciliationMetric.TOTAL_DIVIDENDS_GROSS.value: dividends_gross,
            ReconciliationMetric.TOTAL_DIVIDENDS_TAX.value: dividends_tax,
            ReconciliationMetric.TOTAL_DIVIDENDS_NET.value: dividends_gross + dividends_tax,
            ReconciliationMetric.TOTAL_COUPONS.value: coupon_total,
        }
    )

    for balance in dataset.tables.get("CashBalances", []):
        key = _dimension_key(
            year=int(balance["year"]),
            currency=_text(balance.get("currency")),
        )
        dataset.raw_totals.cash_by_currency[key] = (
            dataset.raw_totals.cash_by_currency.get(key, Decimal("0"))
            + _decimal(balance.get("ending_cash"))
        )

    for report in reports:
        for position in report.positions:
            year = _position_snapshot_year(position, report)
            if year is None:
                continue
            isin = _text(position.get("isin"))
            if not isin:
                continue
            key = _dimension_key(year=year, instrument_key=isin)
            dataset.raw_totals.positions_by_key[key] = (
                dataset.raw_totals.positions_by_key.get(key, Decimal("0"))
                + _decimal(position.get("quantity"))
            )


def _unprocessed_trade(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "severity": "warning",
        "reason": "unsupported_operation",
        "details": f"Unsupported Alatay security operation: {row.get('operation')}",
        "source_sheet": "Trades",
        "source_report": row.get("source_report"),
        "trade_id": f"{Path(str(row.get('source_report'))).name}:{row.get('source_row')}",
        "date_time": row.get("date_time"),
        "symbol": row.get("isin"),
        "isin": row.get("isin"),
        "asset_type": _asset_type(row.get("security_type")),
        "currency": row.get("currency"),
        "quantity": row.get("quantity"),
        "price": row.get("price"),
        "amount": row.get("amount"),
        "commission": row.get("commission"),
    }


def _metadata_value(values: Sequence[str], label: str) -> str | None:
    for index, value in enumerate(values):
        if value == label:
            return values[index + 1] if index + 1 < len(values) else ""
    return None


def _parse_period(value: str) -> tuple[date | None, date | None]:
    if "-" not in value:
        report_date = _parse_date(value)
        return report_date, report_date
    start, end = value.split("-", 1)
    return _parse_date(start), _parse_date(end)


def _parse_amount_currency(value: Any) -> tuple[Decimal, str | None]:
    parts = _clean_text(value).split()
    currency = parts[-1].upper() if parts and len(parts[-1]) == 3 and parts[-1].isalpha() else None
    amount_parts = parts[:-1] if currency else parts
    return _decimal("".join(amount_parts)), currency


def _parse_date(value: Any) -> date | None:
    text = _clean_text(value)
    try:
        return datetime.strptime(text, "%d.%m.%Y").date()
    except ValueError:
        return None


def _decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    text = str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".").strip()
    return Decimal(text or "0")


def _clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").split()).strip()


def _text(value: Any) -> str | None:
    text = _clean_text(value)
    return text or None


def _asset_type(value: Any) -> str:
    normalized = _clean_text(value).casefold()
    if "облигац" in normalized:
        return "Bonds"
    return "Stocks"


def _normalized_exchange(value: Any) -> str | None:
    exchange = _text(value)
    if exchange and "KASE" in exchange.upper():
        return "KASE"
    return exchange


def _country_from_values(raw_country: Any, isin: str | None) -> str | None:
    country = _text(raw_country)
    if country:
        return country.upper()
    return isin[:2].upper() if isin and len(isin) >= 2 else None


def _max_report_year(reports: Sequence[ParsedAlatayReport]) -> int | None:
    return max((report.period_end.year for report in reports if report.period_end), default=None)


def _year_from_datetime(value: Any) -> int | None:
    text = _clean_text(value)
    return int(text[:4]) if len(text) >= 4 and text[:4].isdigit() else None


def _year_from_date(value: Any) -> int | None:
    return _year_from_datetime(value)


def _date_from_datetime(value: Any) -> str | None:
    text = _clean_text(value)
    return text[:10] if len(text) >= 10 else None


def _is_internal_security_transfer(value: Any) -> bool:
    operation = _clean_text(value).casefold()
    return "перевод" in operation and "без смены прав собственности" in operation


def _same_account_id(left: Any, right: Any) -> bool:
    left_text = _clean_text(left)
    right_text = _clean_text(right)
    if left_text.isdigit() and right_text.isdigit():
        return left_text.lstrip("0") == right_text.lstrip("0")
    return left_text == right_text


def _position_snapshot_year(
    position: Mapping[str, Any],
    report: ParsedAlatayReport,
) -> int | None:
    snapshot_text = _text(position.get("_snapshot_date"))
    snapshot_date = date.fromisoformat(snapshot_text) if snapshot_text else report.period_end
    if snapshot_date is None:
        return None
    if position.get("_snapshot_kind") == "opening":
        return snapshot_date.year - 1
    return snapshot_date.year


def _annual_rate(
    fx_provider: AnnualFxRateProvider,
    year: int | None,
    currency: str | None,
    warnings: list[str],
) -> Decimal | None:
    if year is None or not currency:
        return None
    rate = fx_provider.rate(year, currency)
    if rate is None:
        warning = f"Missing annual NBK FX rate for {currency}/{year}; KZT fields left empty."
        if warning not in warnings:
            warnings.append(warning)
    return rate


def _money_text(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _amount_kzt(amount: Decimal, rate: Decimal | None) -> str | None:
    return str(amount * rate) if rate is not None else None


def _dimension_key(
    *,
    metric: str | None = None,
    year: int | None = None,
    currency: str | None = None,
    instrument_key: str | None = None,
) -> str:
    return "|".join("" if value is None else str(value) for value in (metric, year, currency, instrument_key))
