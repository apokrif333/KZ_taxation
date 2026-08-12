"""In-place preparation of an audit workbook's Trades sheet for Form 270.05."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from kztax270.canonical.trade_enrichment import (
    classify_form270_05_sources,
    enrich_trades_with_kzt,
)
from kztax270.canonical.workbook_schema import required_columns
from kztax270.reference.fx import AnnualFxRateProvider

from .audit_workbook import NUMERIC_WORKBOOK_COLUMNS, display_column_name


def prepare_form270_05_trades_workbook(
    workbook_path: Path,
    fx_provider: AnnualFxRateProvider,
) -> Path:
    """Add KZT/source columns, classify purchases, and sort Trades chronologically."""

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("Preparing Form 270.05 workbooks requires openpyxl.") from exc

    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path)
    if "Trades" not in workbook.sheetnames:
        return path

    worksheet = workbook["Trades"]
    columns = list(required_columns("Trades"))
    source_column = "source_of_expense"
    columns.insert(columns.index(source_column) + 1, "cumulative_source_of_expense")
    header_aliases = {display_column_name(column): column for column in columns}
    header_aliases.update({column: column for column in columns})
    source_headers = [cell.value for cell in worksheet[1]]
    canonical_by_index = {
        index: header_aliases.get(str(value), _snake_case(str(value)))
        for index, value in enumerate(source_headers)
        if value not in (None, "")
    }

    trades: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            canonical: values[index] if index < len(values) else None
            for index, canonical in canonical_by_index.items()
        }
        if any(value not in (None, "") for value in row.values()):
            trades.append(row)

    enrich_trades_with_kzt(trades, fx_provider)
    trades = classify_form270_05_sources(trades)

    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append([display_column_name(column) for column in columns])
    for row in trades:
        worksheet.append(
            [
                _excel_value(
                    row.get(column),
                    numeric=column in NUMERIC_WORKBOOK_COLUMNS or column == "cumulative_source_of_expense",
                )
                for column in columns
            ]
        )
    workbook.save(path)
    return path


def _excel_value(value: Any, *, numeric: bool) -> Any:
    if not numeric or value in (None, ""):
        return value
    try:
        number = Decimal(str(value))
    except Exception:
        return None
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def _snake_case(value: str) -> str:
    return value.strip().replace(" ", "_").replace("-", "_").casefold()
