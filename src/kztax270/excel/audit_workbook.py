"""Canonical Excel audit workbook writer."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping, Sequence

from kztax270.canonical.schema import CanonicalDataset
from kztax270.canonical.workbook_schema import CANONICAL_WORKBOOK_SHEETS, YEARS_RESULTS_TABLE_COLUMNS


NUMERIC_WORKBOOK_COLUMNS = {
    "year",
    "multiplier",
    "strike",
    "quantity",
    "proceeds",
    "value",
    "realized_pl",
    "gross_amount",
    "withholding_tax",
    "net_amount",
    "kzt_rate",
    "gross_amount_kzt",
    "withholding_tax_kzt",
    "net_amount_kzt",
    "tax",
    "tax_kzt",
    "price",
    "amount",
    "commission",
    "amount_with_commission",
    "enter_quantity",
    "enter_price",
    "enter_multiplier",
    "enter_amount",
    "enter_commission",
    "exit_quantity",
    "exit_price",
    "exit_multiplier",
    "exit_amount",
    "exit_commission",
    "acquisition_cost_with_commission",
    "pnl_before_commission",
    "pnl_after_all_commissions",
    "pnl",
    "exit_amount_kzt",
    "acquisition_cost_with_commission_kzt",
    "pnl_before_commission_kzt",
    "pnl_after_all_commissions_kzt",
    "pnl_kzt",
    "amount_kzt",
    "ending_cash",
    "ending_cash_kzt",
    "only_profit",
    "only_profit_kzt",
    "withhold_kzt",
    "tax_kzt_withhold",
    "broker_value",
    "canonical_value",
    "difference",
    "tolerance",
}

YEARS_RESULTS_DIMENSION_COLUMNS = frozenset({"year", "flag", "country", "tax_exchange", "currency"})
YEARS_RESULTS_KZT_COLUMNS = frozenset({"pnl_kzt", "amount_kzt", "only_profit_kzt", "withhold_kzt", "tax_kzt", "tax_kzt_withhold"})
YEARS_RESULTS_AMOUNT_FORMAT = "0.00"
YEARS_RESULTS_KZT_FORMAT = '#,##0.00 "₸"'


class ExcelAuditWorkbookWriter:
    """Write one canonical audit workbook per broker account."""

    def write(self, dataset: CanonicalDataset, output_path: Path) -> Path:
        try:
            import pandas as pd  # type: ignore
        except Exception as exc:
            raise RuntimeError("Excel writing requires pandas and openpyxl. Install project ETL dependencies.") from exc

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet in CANONICAL_WORKBOOK_SHEETS:
                records = dataset.tables.get(sheet.name, [])
                if sheet.name == "Years_Results":
                    write_years_results_sheet(writer, records)
                    continue
                if sheet.name == "CashBalances":
                    records = [
                        {
                            **record,
                            "broker": record.get("broker") or dataset.metadata.broker,
                            "account_id": record.get("account_id") or dataset.metadata.account_id,
                        }
                        for record in records
                    ]
                df = pd.DataFrame(records)
                df = ensure_columns(df, sheet.required_columns)
                df = coerce_numeric_columns_for_excel(df)
                df = df.rename(columns=display_column_name)
                df.to_excel(writer, sheet_name=sheet.name, index=False)
        return output_path


def ensure_columns(df: Any, required_columns: Sequence[str]) -> Any:
    for column in required_columns:
        if column not in df.columns:
            df[column] = None
    return df[list(required_columns)]


def coerce_numeric_columns_for_excel(df: Any) -> Any:
    """Write numeric canonical fields as Excel numbers, not text cells."""

    import pandas as pd  # type: ignore

    df = df.copy()
    for column in df.columns:
        if column in NUMERIC_WORKBOOK_COLUMNS:
            df[column] = pd.to_numeric(df[column], errors="coerce")
    return df


def display_column_name(column: str) -> str:
    explicit = {
        "id": "ID",
        "isin": "ISIN",
        "figi": "FIGI",
        "cusip": "CUSIP",
        "pnl": "PnL",
        "pnl_kzt": "PnL_KZT",
        "kzt": "KZT",
        "usd": "USD",
        "kzt_rate": "KZT_Rate",
        "pnl_before_commission": "PnL_Before_Commission",
        "pnl_after_all_commissions": "PnL_After_All_Commissions",
        "pnl_kzt": "PnL_KZT",
        "tax_kzt": "Tax_KZT",
        "tax_kzt_withhold": "Tax_KZT_Withhold",
        "withhold_kzt": "Withhold_KZT",
        "amount_kzt": "Amount_KZT",
        "only_profit": "OnlyProfit",
        "only_profit_kzt": "OnlyProfit_KZT",
        "fx": "FX",
    }
    if column in explicit:
        return explicit[column]
    parts = column.split("_")
    return "_".join(explicit.get(part, part[:1].upper() + part[1:]) for part in parts)


def write_years_results_sheet(writer: Any, records: Sequence[Mapping[str, Any]]) -> None:
    import pandas as pd  # type: ignore
    grouped: dict[str, list[Mapping[str, Any]]] = {}
    for record in records:
        table_name = str(record.get("table") or "Unclassified")
        grouped.setdefault(table_name, []).append(record)

    row_idx = 0
    ordered_table_names = list(YEARS_RESULTS_TABLE_COLUMNS)
    ordered_table_names.extend(name for name in grouped if name not in YEARS_RESULTS_TABLE_COLUMNS)
    for table_name in ordered_table_names:
        table_records = grouped.get(table_name, [])
        if not table_records:
            continue
        title_row = row_idx
        columns = YEARS_RESULTS_TABLE_COLUMNS.get(table_name, tuple(key for key in table_records[0] if key != "table"))
        table_records = sorted(
            table_records,
            key=lambda row: (
                -1 if row.get("year") in (None, "") else int(row.get("year")),
                str(row.get("flag") or ""),
                str(row.get("tax_exchange") or row.get("exchange") or ""),
                str(row.get("currency") or ""),
            ),
        )

        pd.DataFrame([[table_name]]).to_excel(
            writer,
            sheet_name="Years_Results",
            startrow=row_idx,
            index=False,
            header=False,
        )
        row_idx += 1
        df = pd.DataFrame(table_records)
        df = ensure_columns(df, columns)
        df = coerce_numeric_columns_for_excel(df)
        df = df.rename(columns=display_column_name)
        df.to_excel(writer, sheet_name="Years_Results", startrow=row_idx, index=False)
        _style_years_results_block(
            writer,
            title_row=title_row,
            table_name=table_name,
            columns=columns,
            record_count=len(table_records),
        )
        row_idx += len(table_records) + 3


def _style_years_results_block(
    writer: Any,
    *,
    title_row: int,
    table_name: str,
    columns: Sequence[str],
    record_count: int,
) -> None:
    """Apply the visual hierarchy used by the yearly result blocks."""

    from openpyxl.styles import Alignment, Border, Font, Side  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    ws = writer.sheets["Years_Results"]
    title_excel_row = title_row + 1
    header_excel_row = title_excel_row + 1
    last_column = len(columns)
    last_column_letter = get_column_letter(last_column)
    black = Side(style="thin", color="000000")
    all_sides = Border(left=black, right=black, top=black, bottom=black)

    # The table title spans exactly the width of its columns.
    for column_idx in range(1, last_column + 1):
        cell = ws.cell(row=title_excel_row, column=column_idx)
        cell.border = Border(
            left=black if column_idx == 1 else Side(style=None),
            right=black if column_idx == last_column else Side(style=None),
            top=black,
            bottom=black,
        )
    ws.merge_cells(f"A{title_excel_row}:{last_column_letter}{title_excel_row}")
    title_cell = ws.cell(row=title_excel_row, column=1)
    title_cell.value = table_name
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_idx in range(1, last_column + 1):
        header = ws.cell(row=header_excel_row, column=column_idx)
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal="center", vertical="center")
        header.border = all_sides

    data_start_row = header_excel_row + 1
    data_end_row = data_start_row + record_count - 1
    for column_idx, column_name in enumerate(columns, start=1):
        for row_idx in range(data_start_row, data_end_row + 1):
            cell = ws.cell(row=row_idx, column=column_idx)
            if column_name in YEARS_RESULTS_DIMENSION_COLUMNS:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = all_sides
            else:
                cell.number_format = (
                    YEARS_RESULTS_KZT_FORMAT if column_name in YEARS_RESULTS_KZT_COLUMNS else YEARS_RESULTS_AMOUNT_FORMAT
                )


def table_records_for_workbook(
    tables: Mapping[str, Sequence[Mapping[str, Any]]]
) -> dict[str, list[dict[str, Any]]]:
    return {sheet.name: [dict(row) for row in tables.get(sheet.name, [])] for sheet in CANONICAL_WORKBOOK_SHEETS}
